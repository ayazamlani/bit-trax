import datetime
import urllib.request
import bs4 as bs
import json
import openpyxl
from openpyxl import Workbook
from bittrex import Bittrex
from openpyxl.chart import (
    PieChart,
    Reference
)
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font

# Change this to your API data. For your security, only enable API key to read data. Not to buy and sell orders.
API_KEY = 'INSERT_API_KEY_HERE'
API_SECRET = 'INSERT_API_SECRET_HERE'


# Static global data Do Not Change
API = Bittrex(API_KEY, API_SECRET)
url = ('https://api.coinmarketcap.com/v1/ticker/')

# Global Variables
currency_data = []
current_balance = 0


def main():
    # main method
    retrieveBittrexData()
    getCurrencyPrices()
    writePricesToExcel()


def retrieveBittrexData():
    # function to collect Symbols and number of shares owned on bittrex
    data = API.get_balances()
    for item in data['result']:
        if item['Balance'] >= .01:
            currency = {}
            currency['Symbol'] = item['Currency']
            currency['Balance'] = item['Balance']
            currency_data.append(currency)


def getCurrencyPrices():
    # function to check prices if stocks owned on bittrex are on coin market cap
    # Gathers current stock price and calculates the total pricing of shares
    data = urllib.request.urlopen(url).read()
    soup = bs.BeautifulSoup(data, 'lxml')
    body = soup.find('p')
    currency_price_data = json.loads(str(body.contents[0]))
    total_balance = 0
    for item in currency_price_data:
        for d in currency_data:
            if item['symbol'] == d['Symbol']:
                symbol = item['symbol']
                price = float(item['price_usd'])
                d['Price'] = price
                balance = d['Balance']
                share_total = price * balance
                d['Total'] = share_total
                total_balance += share_total
    global current_balance
    current_balance = total_balance


def writePricesToExcel():
    # populates data in to an excel sheet.
    count = 1
    try:
        wb = openpyxl.load_workbook('Bittrex_Data.xlsx', data_only=True)
    except:
        wb = Workbook()
    sheet_name = str(datetime.datetime.now())[:10] + ' %d' % count
    while sheet_name in wb.sheetnames:
        count += 1
        sheet_name = str(datetime.datetime.now())[:10] + ' %d' % count
    wb.create_sheet(title=sheet_name)
    ws = wb.get_sheet_by_name(sheet_name)
    ws['K1'] = 'Date'
    ws['L1'] = 'Symbol'
    ws['M1'] = 'Price (USD)'
    ws['N1'] = '# of Shares Owned'
    ws['O1'] = 'Value of Shares'
    for i, item in enumerate(currency_data):
        ws.cell(row=i + 2, column=11).value = datetime.datetime.now()
        ws.cell(row=i + 2, column=12).value = item['Symbol']
        ws.cell(row=i + 2, column=13).value = item['Price']
        ws.cell(row=i + 2, column=13).number_format = '$0.0000'
        ws.cell(row=i + 2, column=14).value = item['Balance']
        ws.cell(row=i + 2, column=15).value = item['Total']
        ws.cell(row=i + 2, column=15).number_format = '$00.00'
    ws['P1'] = 'Portfolio Total Value'
    ws['P1'].font = Font(bold=True)
    ws['P2'] = current_balance
    ws['P2'].font = Font(bold=True)
    ws['P2'].number_format = '$0,000.00'
    if 'Sheet' in wb.sheetnames:
        default_sheet = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(default_sheet)
    # analyzeData(wb)
    wb.save('Bittrex_Data.xlsx')
    for sheet in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(sheet)
        create_Pie_Chart(ws)
    wb.save('Bittrex_Data.xlsx')


def create_Pie_Chart(ws):
    pie = PieChart()
    labels = Reference(ws, min_col=12, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=15, min_row=1, max_row=ws.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Portfolio"
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.separator = "\n"
    pie.width = 20
    pie.height = 20
    ws.add_chart(pie, "A1")


main()
